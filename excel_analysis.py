import pandas as pd
pd.set_option('display.float_format', '{:.2f}'.format)

import numpy as np
import openpyxl


#-------------------------------------------------------------------------------------------
# reading data from excel file
data_file = 'Sales Records.xlsx'

workbook = openpyxl.load_workbook('data//'+data_file)
sheets = workbook.sheetnames

# print excel workbook and sheets that are in the workbook
print(f'> excel file is: ',data_file)
print(f'> excel sheet names: ',sheets)

#-------------------------------------------------------------------------------------------
# get rows and columns for data in 'americas' sheet
americas_sheet = workbook['americas']
americas_data = americas_sheet.values
americas_df = pd.DataFrame(americas_data)
print(americas_df)  

header = americas_df.iloc[0]
americas_df = americas_df[1:]
americas_df.columns = header
print(americas_df.dtypes)

#americas_df['Order Date']= pd.to_datetime(americas_df['Order Date'])
americas_df = americas_df.astype({  'Units Sold': 'int', 
                                    'Unit Price': 'float', 
                                    'Unit Cost': 'float', 
                                    'Total Revenue': 'float', 
                                    'Total Cost': 'float', 
                                    'Total Profit': 'float'})
                                    
                                    
#print(americas_df.dtypes)
americas_shape = americas_df.shape
print(f'> rows: ',americas_shape[0])
print(f'> columns: ',americas_shape[1])


#-------------------------------------------------------------------------------------------
# let's do some calculations
# get grand total of revenue for america
america_rev_total = americas_df['Total Revenue'].sum()
print(f'> total revenue = {america_rev_total}')

# get grand total of cost for america
america_cost_total = americas_df['Total Cost'].sum()
print(f'> total cost = {america_cost_total}')

# get grand total of cost for america
america_profit_total = americas_df['Total Profit'].sum()
print(f'> total profit = {america_profit_total}')

america_profit_total_check = america_rev_total - america_cost_total
print(f'> total profit (check) = {america_profit_total_check}')


#-------------------------------------------------------------------------------------------
# calculate total profit and compare to the total profit field in excel
americas_df['Total Profit calc'] =  americas_df['Total Revenue']-americas_df['Total Cost']
americas_df['Total Profit check']= np.where(abs(americas_df['Total Profit calc']-americas_df['Total Profit'])<=1.0, 'Correct', 'Incorrect!!')

americas_df.loc[(americas_df['Sales Channel'] == 'Online') & (americas_df['Order Priority'].isin(['H','M','C'])), 'Process Priority'] = 'Urgent'
americas_df.loc[(americas_df['Sales Channel'] == 'Online') & (americas_df['Order Priority']=='L'), 'Process Priority'] = 'Medium'
americas_df.loc[americas_df['Sales Channel'] == 'Offline', 'Process Priority'] = 'Low'

americas_df['High Rev Urgent'] = 'No'
americas_df.loc[(americas_df['Process Priority'] == 'Urgent') & (americas_df['Total Revenue'] > 1000000), 'High Rev Urgent'] = 'Yes'

#-------------------------------------------------------------------------------------------
# add grand total of revenue for america and output to excel workbook
americas_df.loc['Total', 'Total Revenue']= round(americas_df['Total Revenue'].sum(), 0)
americas_df.loc['Total', 'Total Cost']= round(americas_df['Total Cost'].sum(), 0)
americas_df.loc['Total', 'Total Profit']= round(americas_df['Total Profit'].sum(), 0)
americas_df.loc['Total', 'Units Sold']= round(americas_df['Units Sold'].sum(), 0)

americas_df.iloc[-1, americas_df.columns.get_loc('Region')] = americas_df.index[-1]

'''
# save updated dataframe to excel file
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "americas"

for r in dataframe_to_rows(americas_df, index=False, header=True):
    ws.append(r)
 
wb.save('data//Sales Records processed.xlsx')
'''


#-------------------------------------------------------------------------------------------
# add average and median for america and output to excel workbook
americas_df.loc['Median', 'Total Revenue']= round(americas_df['Total Revenue'].median(),2)

americas_df.loc['Median', 'Total Cost']= round(americas_df['Total Cost'].median(),2)
americas_df.loc['Median', 'Total Profit']= round(americas_df['Total Profit'].median(),2)
americas_df.loc['Median', 'Units Sold']= round(americas_df['Units Sold'].median(),2)
americas_df.loc['Median', 'Unit Price']= round(americas_df['Unit Price'].median(),2)
americas_df.loc['Median', 'Unit Cost']= round(americas_df['Unit Cost'].median(),2)

#americas_df['Region'].iloc[-1] = americas_df.index[-1]
americas_df.iloc[-1, americas_df.columns.get_loc('Region')] = americas_df.index[-1]


americas_df.loc['Mean', 'Total Revenue']= round(americas_df['Total Revenue'].mean(),2)
americas_df.loc['Mean', 'Total Cost']= round(americas_df['Total Cost'].mean(),2)
americas_df.loc['Mean', 'Total Profit']= round(americas_df['Total Profit'].mean(),2)
americas_df.loc['Mean', 'Units Sold']= round(americas_df['Units Sold'].mean(),2)
americas_df.loc['Mean', 'Unit Price']= round(americas_df['Unit Price'].mean(),2)
americas_df.loc['Mean', 'Unit Cost']= round(americas_df['Unit Cost'].mean(),2)

#americas_df['Region'].iloc[-1] = americas_df.index[-1]
americas_df.iloc[-1, americas_df.columns.get_loc('Region')] = americas_df.index[-1]
print(">>", americas_df)

#-------------------------------------------------------------------------------------------
# save updated dataframe to excel file
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "americas"

for r in dataframe_to_rows(americas_df, index=False, header=True):
    ws.append(r)
 
#wb.save('data//Sales Records processed.xlsx')


#-------------------------------------------------------------------------------------------
# format the header
#black = 'FF000000', white = 'FFFFFFFF', red = 'FFFF0000', blue = 'FF0000FF', green = 'FF00FF00', yellow = 'FFFFFF00'

    
font_black_bold = openpyxl.styles.Font(color='FF000000', bold=True)
font_red_italic = openpyxl.styles.Font(color='FFFF0000', italic=True) 
bd_thick = openpyxl.styles.Side(style='thick', color="FF000000")

col_nums = len(ws[1])
calc_cols = 4
i = 1

for cell in ws[1]:
    cell.border = openpyxl.styles.Border(bottom=bd_thick)
    
    if i > col_nums - calc_cols:
        cell.font = font_red_italic
    else:
        cell.font = font_black_bold
        
    i += 1 

#-------------------------------------------------------------------------------------------
# format the last three rows (total, mean, median)

# get maximum number of rows.
max_row = ws.max_row

# we need to format the last three rows
last_rows = [max_row-2, max_row-1, max_row]

font_red_italic = openpyxl.styles.Font(color='FFFF0000', italic=True)                  
bd_double = openpyxl.styles.Side(style='double', color="FF000000")
bd_thin = openpyxl.styles.Side(style='thin', color="FF000000")

# apply style to the last 3 rows (total, mean, median)
for row in last_rows:
    for cell in ws[row]:
        cell.font = font_red_italic
        
        if row==max_row:
            cell.border = openpyxl.styles.Border(top=bd_thin, bottom=bd_double)

        else:
            cell.border = openpyxl.styles.Border(top=bd_thin, bottom=bd_thin)


#-------------------------------------------------------------------------------------------
# freeze panes
 
ws.freeze_panes = ws['A2']


#-------------------------------------------------------------------------------------------
# lets add another sheet and perform some excel functions

# COUNTIF:  count the number of orders for household items
count_item_household = len(americas_df[americas_df['Item Type']=='Household'])
count_item_household_2 = americas_df[americas_df['Item Type']=='Household'].shape[0]

# SUMIF:  total of revenue for Household items
sum_item_household = americas_df[americas_df['Item Type']=='Household']['Total Revenue'].sum()

# AVERAGEIF:  average revenue for Household items
avg_item_household = round(americas_df[americas_df['Item Type']=='Household']['Total Revenue'].mean(), 0)


#-------------------------------------------------------------------------------------------
# we can use more complex conditions 

# AVERAGEIF:  average revenue for household items ordered with medium priority 
avg_item_hh_prior_m = round(americas_df[(americas_df['Item Type']=='Household') & (americas_df['Order Priority']=='M')]['Total Revenue'].mean(), 0)

# SUMIF:  sum revenue for household items ordered with medium priority 
sum_item_hh_prior_m = americas_df[(americas_df['Item Type']=='Household') & (americas_df['Order Priority']=='M')]['Total Revenue'].sum()

# AVERAGEIF:  average for high revenue orders (greater than $100k) for household items ordered with high priority 
avg_r1m_ihh_ph = round(americas_df[ (americas_df['Total Revenue']>1000000) & 
                                    (americas_df['Item Type']=='Household') &
                                    (americas_df['Order Priority']=='H')]['Total Revenue'].mean(), 0)



#-------------------------------------------------------------------------------------------
# add these to the Excel workbook in a new sheet

# create a new sheet in the Workbook
ws2 = wb.create_sheet('americas2')

ws2.cell(row=1, column=1).value = 'Scenario'
ws2.cell(row=1, column=2).value = 'Value'

# let's use python dictionary to store the calculations and give them definitions
scenario_dict = {   'number of orders for household items': count_item_household, 
                    'sum of revenue for household items': sum_item_household, 
                    'average revenue for household items': avg_item_household,
                    'sum revenue for household items ordered with medium priority': sum_item_hh_prior_m,
                    'average revenue for household items ordered with medium priority': avg_item_hh_prior_m,
                    'average for high revenue orders (> $1m) for household items ordered with high priority': avg_r1m_ihh_ph}

row_id = 2

# let's now use the dictionary to enter data in excel file. 
# we will also do some formating for numeric values

from openpyxl.styles import Font

value_fmt = u'$#,###'
bd_thick = openpyxl.styles.Side(style='thick', color="FF000000")
font_black_bold = openpyxl.styles.Font(color='FF000000', bold=True)

for key, value in scenario_dict.items():
    ws2.cell(row=row_id, column=1).value = key
    ws2.cell(row=row_id, column=2).value = value

    if row_id>2:    
        ws2.cell(row=row_id, column=2).number_format = value_fmt
    
    row_id += 1


for cell in ws2[1]:
    cell.border = openpyxl.styles.Border(bottom=bd_thick)
    cell.font = font_black_bold



#-------------------------------------------------------------------------------------------
# create a summary table that provides average profit per country for one particular item

# let's pick cosmetics and create the table in a new sheet.
ws3 = wb.create_sheet('cosmetics')

cosmetics_df = americas_df[americas_df['Item Type']=='Cosmetics']
grouped_df = cosmetics_df[['Country', 'Total Profit']].groupby('Country', as_index=False)['Total Profit'].mean()
print(grouped_df)

# save the dataframe to the cosmetics sheet
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook

value_fmt = u'$#,###.##'
bd_thick = openpyxl.styles.Side(style='thick', color="FF000000")
font_black_bold = openpyxl.styles.Font(color='FF000000', bold=True)

row_id = 1
for r in dataframe_to_rows(grouped_df, index=False, header=True):
    ws3.append(r)

    # this will apply the $ and comma format for dollar figures
    if row_id>1:    
        ws3.cell(row=row_id, column=2).number_format = value_fmt

    # let's update the second column header to "Average Profit" instead of "Total Profit"
    if row_id==1:    
        ws3.cell(row=row_id, column=2).value = 'Average Profit'
        
        for cell in ws3[row_id]:
            cell.border = openpyxl.styles.Border(bottom=bd_thick)
            cell.font = font_black_bold
    
    row_id += 1

#-------------------------------------------------------------------------------------------
# create multiple summary tables that provide average profit per country for multiple items

# let's first create a new sheet called summary
ws4 = wb.create_sheet('summary')

fruits_df = americas_df[americas_df['Item Type']=='Fruits']
fruits_grouped_df = fruits_df[['Country', 'Total Profit']].groupby('Country', as_index=False)['Total Profit'].mean()
#print(fruits_grouped_df)

vegetables_df = americas_df[americas_df['Item Type']=='Vegetables']
vegetables_grouped_df = vegetables_df[['Country', 'Total Profit']].groupby('Country', as_index=False)['Total Profit'].mean()
#print(vegetables_grouped_df)

meat_df = americas_df[americas_df['Item Type']=='Meat']
meat_grouped_df = meat_df[['Country', 'Total Profit']].groupby('Country', as_index=False)['Total Profit'].mean()
#print(meat_grouped_df)


# save the dataframe to the cosmetics sheet
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook

value_fmt = u'$#,###.##'
bd_thick = openpyxl.styles.Side(style='thick', color="FF000000")
font_black_bold = openpyxl.styles.Font(color='FF000000', bold=True)


row_id = 1
tot_row_id = 1

for r in dataframe_to_rows(fruits_grouped_df, index=False, header=True):
    ws4.append(r)

    # this will apply the $ and comma format for dollar figures
    if row_id>1:    
        ws4.cell(row=tot_row_id, column=2).number_format = value_fmt
    
    # let's update the second column header to "Fruits Average Profit" instead of "Total Profit"
    if row_id==1:
        ws4.cell(row=tot_row_id, column=2).value = 'Fruits Average Profit'
        
        for cell in ws4[tot_row_id]:
            cell.border = openpyxl.styles.Border(bottom=bd_thick)
            cell.font = font_black_bold

    row_id += 1
    tot_row_id += 1


row_id = 1
ws4.append(['', ''])
tot_row_id += 1

for r in dataframe_to_rows(vegetables_grouped_df, index=False, header=True):
    ws4.append(r)

    # this will apply the $ and comma format for dollar figures
    if row_id>1:    
        ws4.cell(row=tot_row_id, column=2).number_format = value_fmt
    
    # let's update the second column header to "Vegetables Average Profit" instead of "Total Profit"
    if row_id==1:
        ws4.cell(row=tot_row_id, column=2).value = 'Vegetables Average Profit'
        
        for cell in ws4[tot_row_id]:
            cell.border = openpyxl.styles.Border(bottom=bd_thick)
            cell.font = font_black_bold

    row_id += 1
    tot_row_id += 1


row_id = 1
ws4.append(['', ''])
tot_row_id += 1

for r in dataframe_to_rows(meat_grouped_df, index=False, header=True):
    ws4.append(r)

    # this will apply the $ and comma format for dollar figures
    if row_id>1:    
        ws4.cell(row=tot_row_id, column=2).number_format = value_fmt
    
    # let's update the second column header to "Meat Average Profit" instead of "Total Profit"
    if row_id==1:
        ws4.cell(row=tot_row_id, column=2).value = 'Meat Average Profit'
        
        for cell in ws4[tot_row_id]:
            cell.border = openpyxl.styles.Border(bottom=bd_thick)
            cell.font = font_black_bold

    row_id += 1
    tot_row_id += 1





#-------------------------------------------------------------------------------------------
# create multiple summary tables with an alternate 

# let's first create a new sheet called summary
ws5 = wb.create_sheet('summary2')

fruits_df = americas_df[americas_df['Item Type']=='Fruits']
fruits_grouped_df = fruits_df[['Country', 'Total Profit']].groupby('Country', as_index=False)['Total Profit'].mean()
print(fruits_grouped_df)

vegetables_df = americas_df[americas_df['Item Type']=='Vegetables']
vegetables_grouped_df = vegetables_df[['Country', 'Total Profit']].groupby('Country', as_index=False)['Total Profit'].mean()
print(vegetables_grouped_df)

meat_df = americas_df[americas_df['Item Type']=='Meat']
meat_grouped_df = meat_df[['Country', 'Total Profit']].groupby('Country', as_index=False)['Total Profit'].mean()
print(meat_grouped_df)


combined_df1 = pd.merge(fruits_grouped_df.rename(columns={'Total Profit':'Fruit Avg Profit'}), vegetables_grouped_df.rename(columns={'Total Profit':'Vegetables Avg Profit'}), how='left', on=['Country'])
combined_df2 = pd.merge(combined_df1, meat_grouped_df.rename(columns={'Total Profit':'Meat Avg Profit'}), how='left', on=['Country']).fillna(0.0001)
print(combined_df2)


value_fmt = u'$#,###.##'
bd_thick = openpyxl.styles.Side(style='thick', color="FF000000")
font_black_bold = openpyxl.styles.Font(color='FF000000', bold=True)

row_id = 1
tot_row_id = 1

for r in dataframe_to_rows(combined_df2, index=False, header=True):
    ws5.append(r)

    # this will apply the $ and comma format for dollar figures
    if row_id>1:    
        ws5.cell(row=tot_row_id, column=2).number_format = value_fmt
        ws5.cell(row=tot_row_id, column=3).number_format = value_fmt
        ws5.cell(row=tot_row_id, column=4).number_format = value_fmt

    if row_id==1:
        
        for cell in ws5[tot_row_id]:
            cell.border = openpyxl.styles.Border(bottom=bd_thick)
            cell.font = font_black_bold
    row_id += 1
    tot_row_id += 1



#-------------------------------------------------------------------------------------------
# create pivot tables 

# let's first create a new sheet called summary
ws6 = wb.create_sheet('pivot')

table = pd.pivot_table(americas_df, values=['Total Revenue'], index=['Country'], columns=['Order Priority'], aggfunc=np.sum, fill_value=0)                             
print(table)

table_df = pd.DataFrame(table.to_records())
table_df.columns = [col.replace("('Total Revenue', ", "Revenue_").replace(")", "").replace("'", "") for col in table_df.columns]
print(table_df)

value_fmt = u'$#,###'
bd_thick = openpyxl.styles.Side(style='thick', color="FF000000")
font_black_bold = openpyxl.styles.Font(color='FF000000', bold=True)

row_id = 1
tot_row_id = 1

for r in dataframe_to_rows(table_df, index=False, header=True):
    ws6.append(r)

    # this will apply the $ and comma format for dollar figures
    if row_id>1:    
        ws6.cell(row=tot_row_id, column=2).number_format = value_fmt
        ws6.cell(row=tot_row_id, column=3).number_format = value_fmt
        ws6.cell(row=tot_row_id, column=4).number_format = value_fmt
        ws6.cell(row=tot_row_id, column=5).number_format = value_fmt

    if row_id==1:
        for cell in ws6[tot_row_id]:
            cell.border = openpyxl.styles.Border(bottom=bd_thick)
            cell.font = font_black_bold
    row_id += 1
    tot_row_id += 1


#-------------------------------------------------------------------------------------------
# more pivot tables 

# let's first create a new sheet called summary
ws7 = wb.create_sheet('pivot_2')

table2 = pd.pivot_table(americas_df, values=['Total Revenue'], index=['Country', 'Sales Channel'], columns=['Item Type', 'Order Priority'], aggfunc=np.sum, fill_value=0)                             
print(table2)

table2_df = pd.DataFrame(table2.to_records()).fillna(0)
table2_df.columns = [col.replace("('Total Revenue', ", "Revenue_").replace(")", "").replace("'", "").replace(", ", "_") for col in table2_df.columns]
print(table2_df)

value_fmt = u'$#,###'
bd_thick = openpyxl.styles.Side(style='thick', color="FF000000")
font_black_bold = openpyxl.styles.Font(color='FF000000', bold=True)

row_id = 1
tot_row_id = 1
n_col = len(table2_df.columns)

for r in dataframe_to_rows(table2_df, index=False, header=True):
    ws7.append(r)

    # this will apply the $ and comma format for dollar figures
    if row_id>1:
        for i in range(2, n_col+1):
            ws7.cell(row=tot_row_id, column=i).number_format = value_fmt

    if row_id==1:
        for cell in ws7[tot_row_id]:
            cell.border = openpyxl.styles.Border(bottom=bd_thick)
            cell.font = font_black_bold
    row_id += 1
    tot_row_id += 1



wb.save('data//Sales Records processed.xlsx')